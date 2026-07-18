"""Excel workbook inspect / range read for startup assistant tools."""

from __future__ import annotations

import re
from typing import Any
from uuid import UUID

from sqlalchemy.orm import Session

from app.models.book import Book
from app.services.sources.source_library_service import SourceLibraryService

_XLSX_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm"}


def _is_excel_filename(name: str) -> bool:
    lower = (name or "").lower()
    return any(lower.endswith(s) for s in _XLSX_SUFFIXES)


def _load_workbook_bytes(db: Session, book: Book, source_id: str):
    from io import BytesIO

    from openpyxl import load_workbook

    from app.services.assets.binary_asset_service import BinaryAssetService

    lib = SourceLibraryService(db)
    try:
        sid = UUID(str(source_id))
    except (TypeError, ValueError) as exc:
        raise ValueError("invalid source_id") from exc
    item = lib.get_item(book.id, sid)
    if not item:
        raise ValueError("Source not found")
    filename = item.filename or "workbook.xlsx"
    if not _is_excel_filename(filename):
        raise ValueError(f"not an Excel workbook: {filename}")
    if not item.asset_id:
        raise ValueError("workbook binary missing")
    asset = BinaryAssetService(db).get_asset_for_book(book_id=book.id, asset_id=item.asset_id)
    wb = load_workbook(BytesIO(bytes(asset.content)), read_only=True, data_only=True)
    return wb, filename


def _col_letter(idx: int) -> str:
    """1-based column index to Excel letter."""
    n = idx
    letters = ""
    while n:
        n, rem = divmod(n - 1, 26)
        letters = chr(65 + rem) + letters
    return letters or "A"


def _sheet_profile(ws) -> dict[str, Any]:
    rows_iter = ws.iter_rows(values_only=True)
    headers: list[str] = []
    sample_rows: list[list[Any]] = []
    missing: dict[str, int] = {}
    numeric_ranges: dict[str, list[float]] = {}
    row_count = 0
    col_count = 0
    try:
        first = next(rows_iter)
    except StopIteration:
        return {
            "name": ws.title,
            "rows": 0,
            "columns": 0,
            "headers": [],
            "missing_summary": {},
            "numeric_ranges": {},
            "sample_rows": [],
        }
    headers = [str(c).strip() if c is not None else f"col_{i+1}" for i, c in enumerate(first)]
    col_count = len(headers)
    for h in headers:
        missing[h] = 0
        numeric_ranges[h] = []
    for row in rows_iter:
        row_count += 1
        values = list(row[:col_count]) if row else []
        while len(values) < col_count:
            values.append(None)
        if len(sample_rows) < 5:
            sample_rows.append(
                [None if v is None else (str(v)[:80] if not isinstance(v, (int, float)) else v) for v in values]
            )
        for i, h in enumerate(headers):
            v = values[i] if i < len(values) else None
            if v is None or (isinstance(v, str) and not v.strip()):
                missing[h] = missing.get(h, 0) + 1
                continue
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                numeric_ranges.setdefault(h, []).append(float(v))
    ranges_out: dict[str, list[float]] = {}
    for h, nums in numeric_ranges.items():
        if nums:
            ranges_out[h] = [round(min(nums), 4), round(max(nums), 4)]
    missing_out = {k: v for k, v in missing.items() if v > 0}
    return {
        "name": ws.title,
        "rows": row_count + 1,
        "columns": col_count,
        "headers": headers[:40],
        "missing_summary": missing_out,
        "numeric_ranges": ranges_out,
        "sample_rows": sample_rows,
    }


_RANGE_RE = re.compile(
    r"^\s*([A-Za-z]+)(\d+)\s*:\s*([A-Za-z]+)(\d+)\s*$"
)


def _col_to_index(letter: str) -> int:
    n = 0
    for ch in letter.upper():
        if not ("A" <= ch <= "Z"):
            continue
        n = n * 26 + (ord(ch) - 64)
    return n


def inspect_workbook(db: Session, book: Book, *, source_id: str) -> dict[str, Any]:
    wb, filename = _load_workbook_bytes(db, book, source_id)
    try:
        sheets = [_sheet_profile(wb[name]) for name in wb.sheetnames[:20]]
        return {
            "source_id": source_id,
            "filename": filename,
            "sheet_count": len(wb.sheetnames),
            "sheets": sheets,
        }
    finally:
        wb.close()


def read_sheet_range(
    db: Session,
    book: Book,
    *,
    source_id: str,
    sheet_name: str,
    cell_range: str = "A1:N50",
) -> dict[str, Any]:
    wb, filename = _load_workbook_bytes(db, book, source_id)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"sheet not found: {sheet_name}")
        ws = wb[sheet_name]
        m = _RANGE_RE.match(cell_range or "")
        if not m:
            # default cap
            min_col, min_row, max_col, max_row = 1, 1, 14, 50
        else:
            min_col = _col_to_index(m.group(1))
            min_row = int(m.group(2))
            max_col = _col_to_index(m.group(3))
            max_row = int(m.group(4))
        # Cap cells to avoid context blow-up
        if (max_row - min_row + 1) * (max_col - min_col + 1) > 2000:
            max_row = min_row + max(1, 2000 // max(1, max_col - min_col + 1)) - 1
        rows_out: list[list[Any]] = []
        for row in ws.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
            values_only=True,
        ):
            rows_out.append(
                [
                    None
                    if v is None
                    else (round(float(v), 6) if isinstance(v, float) else (str(v)[:200] if not isinstance(v, (int, bool)) else v))
                    for v in row
                ]
            )
        return {
            "source_id": source_id,
            "filename": filename,
            "sheet_name": sheet_name,
            "cell_range": f"{_col_letter(min_col)}{min_row}:{_col_letter(max_col)}{max_row}",
            "rows": rows_out,
            "row_count": len(rows_out),
        }
    finally:
        wb.close()
